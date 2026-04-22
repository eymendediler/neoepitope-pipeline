import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

INP = "results/mhcflurry_ALL_WITH_MDT.csv"
OUT = "results/mhcflurry_scientific_LONG.xlsx"

KEEP = [
    "tissue",
    "peptide",
    "allele",
    "rank",
    "affinity",
    "presentation_score",
    "gene_name",
    "gene_id",
    "transcript_id",
    "protein_id",
    "start_position",
    "mdt_enst",
    "is_tissue_MDT",
    "top1_median_tpm",
    "delta_median_tpm",
]

def autofit_worksheet(ws, max_width=45):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def add_table(ws, name):
    # add an Excel "Table" with filter dropdowns
    last_row = ws.max_row
    last_col = ws.max_column
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

def main():
    if not os.path.exists(INP):
        raise SystemExit(f"Missing input: {INP}")

    df = pd.read_csv(INP)

    # keep only existing columns
    cols = [c for c in KEEP if c in df.columns]
    df = df[cols].copy()

    # numeric cleanup
    for c in ["rank", "affinity", "presentation_score", "top1_median_tpm", "delta_median_tpm"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # sort: most important first (best rank)
    if "rank" in df.columns:
        df = df.sort_values(["tissue", "rank", "presentation_score"], ascending=[True, True, False])
    else:
        df = df.sort_values(["tissue", "peptide", "allele"])

    # tissue summary
    tissue_summary = pd.DataFrame()
    if "tissue" in df.columns:
        tissue_summary = (
            df.groupby("tissue")
              .agg(
                  rows=("peptide", "size"),
                  unique_peptides=("peptide", "nunique"),
                  unique_genes=("gene_id", "nunique") if "gene_id" in df.columns else ("peptide","nunique"),
                  mdt_available=("mdt_enst", lambda x: int(x.notna().sum())) if "mdt_enst" in df.columns else ("peptide","size"),
                  mdt_match_fraction=("is_tissue_MDT", lambda x: float(pd.Series(x).dropna().mean())) if "is_tissue_MDT" in df.columns else ("peptide","nunique"),
                  best_rank=("rank", "min") if "rank" in df.columns else ("peptide","nunique"),
              )
              .reset_index()
              .sort_values("rows", ascending=False)
        )

    # peptide summary (across tissues+HLA)
    peptide_summary = (
        df.groupby("peptide")
          .agg(
              n_tissues=("tissue", "nunique") if "tissue" in df.columns else ("allele","nunique"),
              tissues=("tissue", lambda x: ", ".join(sorted(set(map(str, x))))) if "tissue" in df.columns else ("allele","nunique"),
              n_hla=("allele", "nunique") if "allele" in df.columns else ("peptide","size"),
              hla_types=("allele", lambda x: ", ".join(sorted(set(map(str, x))))) if "allele" in df.columns else ("peptide","size"),
              best_rank=("rank", "min") if "rank" in df.columns else ("peptide","size"),
              best_affinity=("affinity", "min") if "affinity" in df.columns else ("peptide","size"),
              best_presentation_score=("presentation_score", "max") if "presentation_score" in df.columns else ("peptide","size"),
              any_MDT=("is_tissue_MDT", lambda x: bool(True in set(pd.Series(x).dropna().tolist()))) if "is_tissue_MDT" in df.columns else ("peptide","size"),
          )
          .reset_index()
    )
    if "best_rank" in peptide_summary.columns:
        peptide_summary = peptide_summary.sort_values("best_rank", ascending=True)

    # write xlsx
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="DATA_long", index=False)
        tissue_summary.to_excel(xw, sheet_name="SUMMARY_tissue", index=False)
        peptide_summary.to_excel(xw, sheet_name="SUMMARY_peptide", index=False)

    # formatting
    wb = load_workbook(OUT)

    for sh in ["DATA_long", "SUMMARY_tissue", "SUMMARY_peptide"]:
        ws = wb[sh]
        ws.freeze_panes = "A2"
        # header style
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # add table for filters
        add_table(ws, name=sh.replace("-", "_").replace(" ", "_"))
        # autofit
        autofit_worksheet(ws)

    wb.save(OUT)
    print("Wrote:", OUT)

if __name__ == "__main__":
    main()
